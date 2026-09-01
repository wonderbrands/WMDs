# -*- coding: utf-8 -*-
import json
import csv
import io
import re
import base64
import xlsxwriter
import openpyxl
from odoo import http
from odoo.http import request

class ImportRackeoController(http.Controller):

    def _safe_float(self, val, default=0.0):
        if val is None or val == '':
            return default
        try:
            return float(str(val).strip().replace(',', ''))
        except (ValueError, TypeError):
            return default

    def _format_qty(self, val):
        if val is None or val == '':
            return ''
        try:
            q_float = float(str(val).strip().replace(',', ''))
            if q_float.is_integer():
                return str(int(q_float))
            return str(q_float)
        except (ValueError, TypeError):
            return str(val)

    def _normalize_code(self, code):
        if not code:
            return ""
        return re.sub(r'[^A-Za-z0-9]', '', str(code)).lower()

    def _prefetch_pos(self, po_names):
        po_map = {}
        if not po_names:
            return po_map
        po_list = list(set(filter(None, [str(n).strip() for n in po_names])))
        if not po_list:
            return po_map
        pos = request.env['purchase.order'].sudo().search([('name', 'in', po_list)])
        for po in pos:
            po_map[po.name.strip().lower()] = po
            
        missing = [n for n in po_list if n.strip().lower() not in po_map]
        for m in missing:
            po_found = request.env['purchase.order'].sudo().search([('name', '=ilike', m)], limit=1)
            if not po_found:
                po_found = request.env['purchase.order'].sudo().search([('name', 'ilike', m)], limit=1)
            if po_found:
                po_map[m.strip().lower()] = po_found
        return po_map

    def _prefetch_products(self, skus):
        product_map = {}
        if not skus:
            return product_map
        sku_list = list(set(filter(None, [str(s).strip() for s in skus])))
        if not sku_list:
            return product_map
        
        products = request.env['product.product'].sudo().search([
            '|', ('default_code', 'in', sku_list), ('barcode', 'in', sku_list)
        ])
        for p in products:
            if p.default_code:
                product_map[p.default_code.strip().lower()] = p
                product_map[self._normalize_code(p.default_code)] = p
            if p.barcode:
                product_map[p.barcode.strip().lower()] = p
                product_map[self._normalize_code(p.barcode)] = p
                
        # For remaining skus not matched directly, try fallback normalized lookup
        unmatched = [s for s in sku_list if s.strip().lower() not in product_map and self._normalize_code(s) not in product_map]
        if unmatched:
            for raw_sku in unmatched:
                norm_sku = self._normalize_code(raw_sku)
                p_found = request.env['product.product'].sudo().search([
                    '|', ('default_code', '=ilike', raw_sku), ('barcode', '=ilike', raw_sku)
                ], limit=1)
                if not p_found:
                    tokens = re.findall(r'[A-Za-z]+|\d+', raw_sku)
                    if tokens:
                        pattern = '%'.join(tokens)
                        candidates = request.env['product.product'].sudo().search([
                            '|', ('default_code', 'ilike', pattern), ('barcode', 'ilike', pattern)
                        ], limit=10)
                        for cand in candidates:
                            if self._normalize_code(cand.default_code) == norm_sku or self._normalize_code(cand.barcode) == norm_sku:
                                p_found = cand
                                break
                        if not p_found and candidates:
                            p_found = candidates[0]
                if p_found:
                    product_map[raw_sku.strip().lower()] = p_found
                    product_map[norm_sku] = p_found
                    if p_found.default_code:
                        product_map[p_found.default_code.strip().lower()] = p_found
                    if p_found.barcode:
                        product_map[p_found.barcode.strip().lower()] = p_found
                        
        return product_map

    def _prefetch_locations(self, loc_names):
        loc_map = {}
        if not loc_names:
            return loc_map
        loc_list = list(set(filter(None, [str(ln).strip() for ln in loc_names])))
        if not loc_list:
            return loc_map
        
        locations = request.env['stock.location'].sudo().search([
            '|', ('barcode', 'in', loc_list),
            '|', ('name', 'in', loc_list),
                 ('complete_name', 'in', loc_list)
        ])
        for l in locations:
            if l.barcode:
                loc_map[l.barcode.strip().lower()] = l
            if l.name:
                loc_map[l.name.strip().lower()] = l
            if l.complete_name:
                loc_map[l.complete_name.strip().lower()] = l
                
        unmatched = [ln for ln in loc_list if ln.strip().lower() not in loc_map]
        for raw_loc in unmatched:
            loc_found = request.env['stock.location'].sudo().search([
                '|', ('barcode', '=ilike', raw_loc),
                '|', ('name', '=ilike', raw_loc),
                     ('complete_name', '=ilike', raw_loc)
            ], limit=1)
            if loc_found:
                loc_map[raw_loc.strip().lower()] = loc_found
                if loc_found.barcode:
                    loc_map[loc_found.barcode.strip().lower()] = loc_found
                if loc_found.name:
                    loc_map[loc_found.name.strip().lower()] = loc_found
                    
        return loc_map

    def _get_reception_location(self):
        pt_stor = request.env['stock.picking.type'].sudo().search([('sequence_code', '=', 'STOR')], limit=1)
        if pt_stor and pt_stor.default_location_src_id:
            return pt_stor.default_location_src_id
        loc = request.env['stock.location'].sudo().search([('complete_name', '=', 'WH/Recepcion')], limit=1)
        if not loc:
            loc = request.env['stock.location'].sudo().search([('name', '=', 'Recepcion')], limit=1)
        return loc

    def _get_stock_location(self):
        pt_stor = request.env['stock.picking.type'].sudo().search([('sequence_code', '=', 'STOR')], limit=1)
        if pt_stor and pt_stor.default_location_dest_id:
            return pt_stor.default_location_dest_id
        loc = request.env['stock.location'].sudo().search([('complete_name', '=', 'WH/Stock')], limit=1)
        return loc

    def _validate_and_match_rackeo_rows(self, raw_rows):
        validated_rows = []
        if not raw_rows:
            return validated_rows

        all_po_names = set()
        all_skus = set()
        all_locs = set()

        for r in raw_rows:
            data = r.get('data', {})
            po = data.get('PO', '').strip()
            sku = data.get('SKU', '').strip()
            ubicacion = data.get('UBICACION', '').strip()
            if po:
                all_po_names.add(po)
            if sku:
                all_skus.add(sku)
            if ubicacion:
                all_locs.add(ubicacion)

        po_map = self._prefetch_pos(list(all_po_names))
        product_map = self._prefetch_products(list(all_skus))
        loc_map = self._prefetch_locations(list(all_locs))
        rec_loc = self._get_reception_location()

        # Group rows by PO
        rows_by_po = {}
        for r in raw_rows:
            po_name = r.get('data', {}).get('PO', '').strip()
            rows_by_po.setdefault(po_name, []).append(r)

        # For each PO, determine available reservoir in WH/Recepcion (considering unreserved open STORs)
        for po_name, p_rows in rows_by_po.items():
            po_record = po_map.get(po_name.lower()) if po_name else None
            
            # If PO exists, calculate reservoir for all products in this PO
            reservoir_cache = {}
            if po_record and rec_loc:
                # Find open STORs for this PO
                open_stors = request.env['stock.picking'].sudo().search([
                    ('origin', '=', po_record.name),
                    ('picking_type_id.sequence_code', '=', 'STOR'),
                    ('state', 'not in', ('done', 'cancel'))
                ])
                # Products requested in this PO batch
                po_prod_ids = set()
                for r in p_rows:
                    sku_str = r.get('data', {}).get('SKU', '').strip()
                    prod = product_map.get(sku_str.lower()) or product_map.get(self._normalize_code(sku_str))
                    if prod:
                        po_prod_ids.add(prod.id)
                
                if po_prod_ids:
                    quants = request.env['stock.quant'].sudo().search([
                        ('location_id', '=', rec_loc.id),
                        ('product_id', 'in', list(po_prod_ids))
                    ])
                    for q in quants:
                        # Total stock in WH/Recepcion
                        reservoir_cache[q.product_id.id] = q.quantity
                        
            # Track planned allocations from reservoir per product
            allocated_qty = {}

            for r in p_rows:
                idx = r.get('index', 0)
                orig_row = r.get('original_row', [])
                data = dict(r.get('data', {}))
                is_excluded = r.get('excluded', False)
                
                errors = []
                warnings = []
                
                po_str = data.get('PO', '').strip()
                sku_str = data.get('SKU', '').strip()
                loc_str = data.get('UBICACION', '').strip()
                pzs_str = data.get('PZS', '').strip()

                product_record = None
                dest_loc_record = None
                pzs_val = 0.0

                # 1. Validate PO
                if not po_str:
                    errors.append({'field': 'PO', 'code': 'missing', 'message': 'El campo PO es obligatorio.'})
                elif not po_record:
                    errors.append({'field': 'PO', 'code': 'not_found', 'message': f'La Orden de Compra "{po_str}" no existe en Odoo.'})
                else:
                    data['PO'] = po_record.name

                # 2. Validate SKU / Product
                if not sku_str:
                    errors.append({'field': 'SKU', 'code': 'missing', 'message': 'El campo SKU es obligatorio.'})
                else:
                    product_record = product_map.get(sku_str.lower()) or product_map.get(self._normalize_code(sku_str))
                    if not product_record:
                        errors.append({'field': 'SKU', 'code': 'not_found', 'message': f'El SKU "{sku_str}" no existe en el catálogo de productos.'})
                    else:
                        data['SKU'] = product_record.default_code or product_record.name

                # 3. Validate Quantity (PZS)
                if not pzs_str:
                    errors.append({'field': 'PZS', 'code': 'missing', 'message': 'El campo PZS (cantidad) es obligatorio.'})
                else:
                    try:
                        pzs_val = self._safe_float(pzs_str)
                        if pzs_val <= 0:
                            errors.append({'field': 'PZS', 'code': 'invalid', 'message': 'La cantidad (PZS) debe ser un número mayor a cero.'})
                        else:
                            data['PZS'] = self._format_qty(pzs_val)
                    except Exception:
                        errors.append({'field': 'PZS', 'code': 'invalid', 'message': 'La cantidad (PZS) no es un número válido.'})

                # 4. Validate Location
                if not loc_str:
                    errors.append({'field': 'UBICACION', 'code': 'missing', 'message': 'El campo UBICACIÓN es obligatorio.'})
                else:
                    dest_loc_record = loc_map.get(loc_str.lower())
                    if not dest_loc_record:
                        errors.append({'field': 'UBICACION', 'code': 'not_found', 'message': f'La ubicación "{loc_str}" no existe en Odoo.'})
                    else:
                        data['UBICACION'] = dest_loc_record.barcode or dest_loc_record.name
                        # Check location blocked
                        if hasattr(dest_loc_record, 'is_location_blocked') and dest_loc_record.is_location_blocked():
                            errors.append({
                                'field': 'UBICACION',
                                'code': 'blocked',
                                'message': f'La ubicación "{dest_loc_record.complete_name}" está bloqueada ({dest_loc_record.block_reason or "Sin motivo"}).'
                            })
                        else:
                            # Check occupancy rules: N1 allows same SKU, others must be empty
                            loc_quants = request.env['stock.quant'].sudo().search([
                                ('location_id', '=', dest_loc_record.id),
                                ('quantity', '>', 0)
                            ])
                            if loc_quants:
                                is_n1 = dest_loc_record.name and dest_loc_record.name.upper().endswith('N1')
                                if is_n1:
                                    if product_record:
                                        other_prods = loc_quants.filtered(lambda q: q.product_id.id != product_record.id)
                                        if other_prods:
                                            errors.append({
                                                'field': 'UBICACION',
                                                'code': 'occupied_n1_diff',
                                                'message': f'La ubicación N1 "{dest_loc_record.name}" contiene un SKU diferente ({other_prods[0].product_id.default_code}).'
                                            })
                                else:
                                    errors.append({
                                        'field': 'UBICACION',
                                        'code': 'not_empty',
                                        'message': f'La ubicación "{dest_loc_record.complete_name}" no está vacía (contiene {loc_quants[0].product_id.default_code}: {loc_quants[0].quantity} pzs).'
                                    })

                # 5. Validate Reservoir in WH/Recepcion
                if product_record and pzs_val > 0 and po_record:
                    total_rec_stock = reservoir_cache.get(product_record.id, 0.0)
                    already_allocated = allocated_qty.get(product_record.id, 0.0)
                    if already_allocated + pzs_val > total_rec_stock:
                        errors.append({
                            'field': 'PZS',
                            'code': 'no_reservoir',
                            'message': f'Falta de reserva de producto en WH/Recepcion para SKU {product_record.default_code}. Requerido acumulado: {already_allocated + pzs_val}, Disponible en recepción: {total_rec_stock}.'
                        })
                    else:
                        allocated_qty[product_record.id] = already_allocated + pzs_val

                val_row = {
                    'index': idx,
                    'original_row': orig_row,
                    'data': data,
                    'excluded': is_excluded,
                    'errors': errors,
                    'warnings': warnings,
                    'po_id': po_record.id if po_record else False,
                    'product_id': product_record.id if product_record else False,
                    'location_id': dest_loc_record.id if dest_loc_record else False,
                }
                validated_rows.append(val_row)

        return validated_rows

    @http.route('/wmds/v2/import_rackeo/validate_file', type='http', auth='user', methods=['POST'], csrf=False)
    def validate_file(self, **post):
        file = post.get('file')
        has_header_str = post.get('has_header', 'true')
        has_header = has_header_str.lower() == 'true'
        column_mapping_str = post.get('column_mapping')
        
        if not file:
            return request.make_response(json.dumps({'error': True, 'error_msg': 'No se cargó ningún archivo.'}), headers=[('Content-Type', 'application/json')])
        
        filename = file.filename or ''
        ext = filename.split('.')[-1].lower() if '.' in filename else ''
        if ext not in ['csv', 'xlsx', 'xls']:
            return request.make_response(json.dumps({'error': True, 'error_msg': 'Formato de archivo no soportado. Debe ser CSV o XLSX.'}), headers=[('Content-Type', 'application/json')])
            
        file_content = file.read()
        if len(file_content) > 1024 * 1024 * 1024:
            return request.make_response(json.dumps({'error': True, 'error_msg': 'El archivo excede el tamaño máximo permitido.'}), headers=[('Content-Type', 'application/json')])
        
        try:
            if ext == 'csv':
                try:
                    csv_text = file_content.decode('utf-8')
                except UnicodeDecodeError:
                    csv_text = file_content.decode('latin1')
                
                # Detect delimiter
                sample = csv_text[:2048]
                delimiter = '\t' if '\t' in sample else (';' if ';' in sample else ',')
                reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
                raw_rows = [list(row) for row in reader]
            else:
                wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
                sheet = wb.active
                raw_rows = []
                for row in sheet.iter_rows(values_only=True):
                    raw_rows.append(list(row))
                wb.close()
        except Exception as e:
            return request.make_response(json.dumps({'error': True, 'error_msg': f'Error al leer el archivo: {str(e)}'}), headers=[('Content-Type', 'application/json')])

        if not raw_rows:
            return request.make_response(json.dumps({'error': True, 'error_msg': 'El archivo está vacío o no contiene filas.'}), headers=[('Content-Type', 'application/json')])

        if has_header:
            header_row = raw_rows[0]
            headers = [str(col).strip() if col is not None else '' for col in header_row]
            data_rows_raw = raw_rows[1:]
        else:
            num_cols = max(len(r) for r in raw_rows) if raw_rows else 0
            headers = [f"Columna {i+1}" for i in range(num_cols)]
            data_rows_raw = raw_rows
            
        data_rows = []
        num_cols = len(headers)
        for row in data_rows_raw:
            formatted_row = []
            for i in range(num_cols):
                if i < len(row):
                    val = row[i]
                    if isinstance(val, float) and val.is_integer():
                        val = int(val)
                    formatted_row.append(str(val).strip() if val is not None else '')
                else:
                    formatted_row.append('')
            data_rows.append(formatted_row)
        
        data_rows = [r for r in data_rows if any(val.strip() for val in r)]
        if not data_rows:
            return request.make_response(json.dumps({'error': True, 'error_msg': 'El archivo no contiene filas con datos.'}), headers=[('Content-Type', 'application/json')])
        
        SYNONYMS = {
            'PO': ['po', 'orden de compra', 'compra', 'purchase_order', 'purchase order', 'orden compra', 'pedido compra', 'order'],
            'SKU': ['sku', 'producto', 'product', 'codigo', 'código', 'default_code', 'referencia', 'articulo', 'artículo'],
            'UBICACION': ['ubicacion', 'ubicación', 'posicion', 'posición', 'location', 'destiny', 'destino', 'estanteria', 'shelf'],
            'PZS': ['pzs', 'piezas', 'cantidad', 'unidades', 'qty', 'count', 'cant', 'units', 'piezas a rackear']
        }
        
        auto_mapping = {}
        if column_mapping_str:
            try:
                auto_mapping = json.loads(column_mapping_str)
            except Exception:
                pass
                
        if not auto_mapping and has_header:
            for key, syn_list in SYNONYMS.items():
                for idx, h in enumerate(headers):
                    h_clean = h.lower().strip()
                    if h_clean == key.lower() or h_clean in [s.lower().strip() for s in syn_list]:
                        auto_mapping[key] = idx
                        break
            
            for key, syn_list in SYNONYMS.items():
                if key not in auto_mapping:
                    for idx, h in enumerate(headers):
                        if idx in auto_mapping.values():
                            continue
                        h_clean = h.lower().strip()
                        if any(syn in h_clean for syn in syn_list):
                            auto_mapping[key] = idx
                            break
                        
        rows_to_validate = []
        for row_idx, row in enumerate(data_rows):
            row_data = {}
            for key in ['PO', 'SKU', 'UBICACION', 'PZS']:
                col_idx = auto_mapping.get(key)
                if col_idx is not None and int(col_idx) < len(row):
                    val = str(row[int(col_idx)]).strip()
                    row_data[key] = val
                else:
                    row_data[key] = ''
            
            rows_to_validate.append({
                'index': row_idx,
                'original_row': row,
                'data': row_data,
                'excluded': False
            })
            
        mapped_results = self._validate_and_match_rackeo_rows(rows_to_validate)
            
        result = {
            'status': 'ok',
            'headers': headers,
            'mapping': auto_mapping,
            'rows': mapped_results
        }
        return request.make_response(json.dumps(result), headers=[('Content-Type', 'application/json')])

    @http.route('/wmds/v2/import_rackeo/validate_rows', type='json', auth='user', methods=['POST'], csrf=True)
    def validate_rows(self, **kw):
        rows = kw.get('rows', [])
        if not rows:
            return {'status': 'ok', 'rows': []}
            
        validated_rows = self._validate_and_match_rackeo_rows(rows)
        return {
            'status': 'ok',
            'rows': validated_rows
        }

    @http.route('/wmds/v2/import_rackeo/process', type='json', auth='user', methods=['POST'], csrf=True)
    def process_rackeo(self, **kw):
        rows = kw.get('rows', [])
        if not rows:
            return {'error': True, 'error_msg': 'No hay datos para procesar.'}
            
        validated_rows = self._validate_and_match_rackeo_rows(rows)
        
        valid_rows = [r for r in validated_rows if not r.get('excluded') and not r.get('errors')]
        if not valid_rows:
            return {'error': True, 'error_msg': 'No hay filas válidas para procesar.'}

        pt_stor = request.env['stock.picking.type'].sudo().search([('sequence_code', '=', 'STOR')], limit=1)
        if not pt_stor:
            return {'error': True, 'error_msg': 'No se encontró el tipo de operación Storage (STOR) en Odoo.'}

        rec_loc = self._get_reception_location()
        dest_stock = self._get_stock_location()

        rows_by_po = {}
        for r in valid_rows:
            po_name = r.get('data', {}).get('PO', '').strip()
            rows_by_po.setdefault(po_name, []).append(r)

        created_stors = []
        row_results_map = {} # map index -> { 'stor_name': ..., 'ok': True/False, 'msg': ... }

        for po_name, po_rows in rows_by_po.items():
            po_record = request.env['purchase.order'].sudo().search([('name', '=', po_name)], limit=1)
            if not po_record:
                for r in po_rows:
                    row_results_map[r['index']] = {'ok': False, 'msg': f'PO {po_name} no encontrada', 'stor': ''}
                continue

            try:
                # 1. Unreserve existing open STORs for this PO to free up WH/Recepcion
                open_stors = request.env['stock.picking'].sudo().search([
                    ('origin', '=', po_record.name),
                    ('picking_type_id.sequence_code', '=', 'STOR'),
                    ('state', 'not in', ('done', 'cancel'))
                ])
                for s in open_stors:
                    s.do_unreserve()

                # 2. Group items by product to create moves
                items_by_prod = {}
                for r in po_rows:
                    prod_id = r['product_id']
                    pzs = self._safe_float(r['data'].get('PZS'))
                    items_by_prod.setdefault(prod_id, []).append({
                        'location_id': r['location_id'],
                        'quantity': pzs,
                        'row_index': r['index']
                    })

                # 3. Create new STOR picking
                new_stor = request.env['stock.picking'].sudo().create({
                    'picking_type_id': pt_stor.id,
                    'location_id': rec_loc.id,
                    'location_dest_id': dest_stock.id,
                    'origin': po_record.name,
                    'user_id': request.env.user.id,
                })

                # 4. Create stock.move for each product
                created_moves = {}
                for prod_id, items in items_by_prod.items():
                    prod = request.env['product.product'].sudo().browse(prod_id)
                    total_prod_qty = sum(item['quantity'] for item in items)
                    
                    move = request.env['stock.move'].sudo().create({
                        'name': f"STOR: {prod.display_name}",
                        'product_id': prod.id,
                        'product_uom_qty': total_prod_qty,
                        'product_uom': prod.uom_id.id,
                        'picking_id': new_stor.id,
                        'location_id': rec_loc.id,
                        'location_dest_id': dest_stock.id,
                    })
                    created_moves[prod_id] = move

                new_stor.action_confirm()
                if new_stor.move_line_ids:
                    new_stor.move_line_ids.unlink()

                # 5. Create stock.move.line for each line
                for prod_id, items in items_by_prod.items():
                    prod = request.env['product.product'].sudo().browse(prod_id)
                    move = created_moves[prod_id]
                    for item in items:
                        request.env['stock.move.line'].sudo().create({
                            'picking_id': new_stor.id,
                            'move_id': move.id,
                            'product_id': prod.id,
                            'product_uom_id': prod.uom_id.id,
                            'location_id': rec_loc.id,
                            'location_dest_id': item['location_id'],
                            'quantity': item['quantity'],
                        })

                # 6. Validate the new STOR picking
                new_stor.button_validate()

                # 7. Adjust remaining demand on previous open STORs
                processed_totals = {prod_id: sum(item['quantity'] for item in items) for prod_id, items in items_by_prod.items()}
                for s in open_stors:
                    for m in s.move_ids:
                        p_id = m.product_id.id
                        if p_id in processed_totals and processed_totals[p_id] > 0:
                            sub_qty = min(m.product_uom_qty, processed_totals[p_id])
                            remaining_demand = m.product_uom_qty - sub_qty
                            processed_totals[p_id] -= sub_qty
                            if remaining_demand > 0:
                                m.write({'product_uom_qty': remaining_demand})
                            else:
                                m._action_cancel()
                    active_moves = s.move_ids.filtered(lambda m: m.state != 'cancel')
                    if not active_moves:
                        s.action_cancel()
                    else:
                        s.action_assign()

                # 8. Create WMDs logs
                total_pzs = sum(self._safe_float(r['data'].get('PZS')) for r in po_rows)
                log_msg = f"Rackeo masivo completado: {new_stor.name} ({len(po_rows)} líneas, {total_pzs} pzs) para PO {po_record.name}."
                request.env['wmds.log'].sudo().create({
                    'purchase': po_record.id,
                    'pick': new_stor.id,
                    'log': log_msg,
                    'user': request.env.user.id,
                })

                created_stors.append({
                    'po': po_record.name,
                    'stor_name': new_stor.name,
                    'stor_id': new_stor.id,
                    'lines_count': len(po_rows),
                    'total_pzs': total_pzs
                })

                for r in po_rows:
                    row_results_map[r['index']] = {'ok': True, 'msg': 'Rackeo completado exitosamente', 'stor': new_stor.name}

            except Exception as e:
                for r in po_rows:
                    row_results_map[r['index']] = {'ok': False, 'msg': f'Error: {str(e)}', 'stor': ''}

        # Generate feedback Excel report
        output_stream = io.BytesIO()
        workbook = xlsxwriter.Workbook(output_stream, {'in_memory': True})
        worksheet = workbook.add_worksheet('Retroalimentacion_Rackeo')
        
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#0F172A',
            'font_color': '#FFFFFF',
            'border': 1,
            'align': 'center'
        })
        cell_format = workbook.add_format({'border': 1})
        error_format = workbook.add_format({'border': 1, 'font_color': '#DC2626'})
        ok_format = workbook.add_format({'border': 1, 'font_color': '#16A34A', 'bold': True})
        
        original_headers = kw.get('headers', [])
        if not original_headers:
            max_col = max(len(row.get('original_row', [])) for row in validated_rows) if validated_rows else 0
            original_headers = [f"Columna {i+1}" for i in range(max_col)]
            
        new_headers = original_headers + ['Ok', 'STOR', 'Estado', 'Detalle']
        for col_idx, h_text in enumerate(new_headers):
            worksheet.write(0, col_idx, h_text, header_format)
            
        for row_idx, row in enumerate(validated_rows):
            excel_row_num = row_idx + 1
            orig_row = row.get('original_row', [])
            
            for col_idx, val in enumerate(orig_row):
                worksheet.write(excel_row_num, col_idx, val, cell_format)
                
            is_excluded = row.get('excluded', False)
            errors = row.get('errors', [])
            res_info = row_results_map.get(row.get('index'))
            
            if is_excluded:
                ok_val = "Excluido"
                stor_val = ""
                estado_val = "Excluido"
                detalle_val = "Fila excluida por el usuario"
                fmt = cell_format
            elif errors:
                ok_val = "❌"
                stor_val = ""
                estado_val = "Con Errores"
                detalle_val = ", ".join([e.get('message', '') for e in errors])
                fmt = error_format
            elif res_info and res_info.get('ok'):
                ok_val = "✅"
                stor_val = res_info.get('stor', '')
                estado_val = "Realizado"
                detalle_val = res_info.get('msg', 'OK')
                fmt = ok_format
            elif res_info:
                ok_val = "❌"
                stor_val = ""
                estado_val = "Error al procesar"
                detalle_val = res_info.get('msg', 'Error')
                fmt = error_format
            else:
                ok_val = "-"
                stor_val = ""
                estado_val = "No procesado"
                detalle_val = ""
                fmt = cell_format
                
            worksheet.write(excel_row_num, len(orig_row), ok_val, fmt)
            worksheet.write(excel_row_num, len(orig_row) + 1, stor_val, cell_format)
            worksheet.write(excel_row_num, len(orig_row) + 2, estado_val, fmt)
            worksheet.write(excel_row_num, len(orig_row) + 3, detalle_val, cell_format)
            
        workbook.close()
        output_stream.seek(0)
        xlsx_data = output_stream.read()
        xlsx_base64 = base64.b64encode(xlsx_data).decode('utf-8')
        
        had_errors = any(not r.get('ok') for r in row_results_map.values()) or any(len(r.get('errors', [])) > 0 for r in validated_rows if not r.get('excluded'))
        
        return {
            'status': 'ok',
            'message': f"Se crearon y validaron {len(created_stors)} rackeo(s) STOR exitosamente.",
            'xlsx_file': xlsx_base64,
            'filename': 'retroalimentacion_rackeo.xlsx',
            'created_stors': created_stors,
            'had_errors': had_errors
        }
